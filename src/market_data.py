"""
Market data loading.

This module loads the original AIRMM Excel workbooks and exposes the market
inputs required by the refactored Python pipeline.

The original VBA project separates information across workbooks:
- AIRMM-MarketData31Oct2019.xlsx contains market data and yield curves;
- AIRMM-Exercises-Basics.xlsm contains shared utilities, including Calendar.

No synthetic data are generated.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .calendar_utils import load_holidays_from_calendar_sheet
from .config import VALUATION_DATE
from .curves import ZeroCurve, load_ois_curve_from_sheet


SWAPTIONS_PHYSICAL_SHEET = "Swaptions Physical"
IR_YIELD_CURVES_SHEET = "IR Yield Curves"
CALENDAR_SHEET = "Calendar"

DEFAULT_MARKET_WORKBOOK_NAME = "AIRMM-MarketData31Oct2019.xlsx"
DEFAULT_BASICS_WORKBOOK_NAME = "AIRMM-Exercises-Basics.xlsm"


@dataclass
class MarketDataWorkbook:
    """
    Container for the loaded AIRMM market data and auxiliary inputs.
    """

    market_path: Path
    basics_path: Path | None
    market_workbook: Any
    basics_workbook: Any | None
    swaptions_physical: Any
    ir_yield_curves: Any
    calendar: Any | None
    holidays: set
    ois_curve: ZeroCurve


def _get_sheet_case_insensitive(workbook: Any, sheet_name: str) -> Any | None:
    """
    Return a worksheet using case-insensitive matching.
    """
    target = sheet_name.lower().strip()

    for name in workbook.sheetnames:
        if name.lower().strip() == target:
            return workbook[name]

    return None


def _load_workbook_data_only(path: Path) -> Any:
    """
    Load an Excel workbook with formulas evaluated as stored values.

    For .xlsm files, keep_vba is not needed because we only read cell values.
    """
    return load_workbook(
        filename=path,
        read_only=False,
        data_only=True,
    )


def load_airmm_workbook(
    market_file_path: str | Path = DEFAULT_MARKET_WORKBOOK_NAME,
    basics_file_path: str | Path | None = DEFAULT_BASICS_WORKBOOK_NAME,
) -> MarketDataWorkbook:
    """
    Load the AIRMM market data and auxiliary workbook.

    Required market-data workbook sheets:
    - Swaptions Physical
    - IR Yield Curves

    Calendar lookup order:
    1. try Calendar sheet in the market workbook;
    2. if absent, try Calendar sheet in AIRMM-Exercises-Basics.xlsm;
    3. if still absent, raise an error.

    This is stricter than a weekend-only fallback because the VBA project
    relies on the Calendar sheet for TARGET business-day logic.
    """
    market_path = Path(market_file_path)

    if not market_path.exists():
        raise FileNotFoundError(
            f"Market workbook not found: {market_path}. "
            "Place AIRMM-MarketData31Oct2019.xlsx in the project root."
        )

    market_workbook = _load_workbook_data_only(market_path)

    ws_swaptions = _get_sheet_case_insensitive(
        market_workbook,
        SWAPTIONS_PHYSICAL_SHEET,
    )
    ws_curves = _get_sheet_case_insensitive(
        market_workbook,
        IR_YIELD_CURVES_SHEET,
    )

    missing = []

    if ws_swaptions is None:
        missing.append(SWAPTIONS_PHYSICAL_SHEET)

    if ws_curves is None:
        missing.append(IR_YIELD_CURVES_SHEET)

    if missing:
        raise KeyError(f"Missing required market workbook sheets: {missing}")

    ws_calendar = _get_sheet_case_insensitive(
        market_workbook,
        CALENDAR_SHEET,
    )

    basics_path: Path | None = None
    basics_workbook = None

    if ws_calendar is None:
        if basics_file_path is None:
            raise KeyError(
                "Calendar sheet not found in market workbook and no basics "
                "workbook path was provided."
            )

        basics_path = Path(basics_file_path)

        if not basics_path.exists():
            raise FileNotFoundError(
                f"Calendar workbook not found: {basics_path}. "
                "Place AIRMM-Exercises-Basics.xlsm in the project root."
            )

        basics_workbook = _load_workbook_data_only(basics_path)

        ws_calendar = _get_sheet_case_insensitive(
            basics_workbook,
            CALENDAR_SHEET,
        )

        if ws_calendar is None:
            raise KeyError(
                f"Calendar sheet not found in {basics_path}."
            )

    holidays = load_holidays_from_calendar_sheet(ws_calendar)
    ois_curve = load_ois_curve_from_sheet(ws_curves)

    return MarketDataWorkbook(
        market_path=market_path,
        basics_path=basics_path,
        market_workbook=market_workbook,
        basics_workbook=basics_workbook,
        swaptions_physical=ws_swaptions,
        ir_yield_curves=ws_curves,
        calendar=ws_calendar,
        holidays=holidays,
        ois_curve=ois_curve,
    )


def workbook_summary(market_data: MarketDataWorkbook) -> dict[str, object]:
    """
    Return a compact summary useful for smoke tests.
    """
    return {
        "market_path": str(market_data.market_path),
        "basics_path": str(market_data.basics_path)
        if market_data.basics_path is not None
        else None,
        "valuation_date": VALUATION_DATE.isoformat(),
        "has_calendar_sheet": market_data.calendar is not None,
        "n_holidays": len(market_data.holidays),
        "ois_curve_nodes": len(market_data.ois_curve.times),
        "first_ois_time": float(market_data.ois_curve.times[0]),
        "last_ois_time": float(market_data.ois_curve.times[-1]),
        "swaptions_physical_shape": (
            market_data.swaptions_physical.max_row,
            market_data.swaptions_physical.max_column,
        ),
    }